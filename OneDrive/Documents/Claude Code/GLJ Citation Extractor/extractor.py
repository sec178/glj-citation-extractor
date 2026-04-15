"""
GLJ Citation Extractor — core extraction logic.

Pipeline:
  1. Extract footnotes from a .docx or .pdf file
  2. Split each footnote on: semicolons, "id.", and "see also"
  3. Return a flat list of citation strings for Excel export
"""

import re
import io

import pandas as pd

_W = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'

# ---------------------------------------------------------------------------
# Step 1: Document extraction
# ---------------------------------------------------------------------------

def extract_footnotes_docx(file_bytes: bytes) -> list[tuple[int, str]]:
    """Return [(footnote_number, text), ...] from a .docx file."""
    import zipfile
    from xml.etree import ElementTree as ET

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            if 'word/footnotes.xml' not in z.namelist():
                return []
            xml_bytes = z.read('word/footnotes.xml')
    except Exception:
        return []

    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return []

    footnotes = []
    for fn_elem in root.findall(f'.//{{{_W}}}footnote'):
        fn_id_str = fn_elem.get(f'{{{_W}}}id', '')
        try:
            fn_id = int(fn_id_str)
        except ValueError:
            continue
        if fn_id <= 0:
            continue

        paragraphs = fn_elem.findall(f'.//{{{_W}}}p')
        text_parts = []
        for p in paragraphs:
            for t in p.findall(f'.//{{{_W}}}t'):
                text_parts.append(t.text or '')
        text = ' '.join(''.join(text_parts).split())
        if text:
            footnotes.append((fn_id, text))

    return footnotes


def extract_footnotes_pdf(file_bytes: bytes) -> list[tuple[int, str]]:
    """
    Extract footnotes from a PDF by scanning the bottom ~38% of each page
    for lines that begin with a footnote number.
    """
    import pdfplumber

    footnote_map: dict[int, str] = {}
    current_fn: int | None = None
    current_text: list[str] = []
    max_fn_seen: int = 0
    footnote_start_re = re.compile(r'^(\d{1,4})[\.\)]\s+\S')

    def flush(fn_id, text_parts):
        if fn_id is not None and text_parts:
            combined = ' '.join(text_parts).strip()
            footnote_map[fn_id] = (footnote_map.get(fn_id, '') + ' ' + combined).strip()

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            height = page.height
            region = page.within_bbox((0, height * 0.62, page.width, height))
            text = region.extract_text() or ''
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                m = footnote_start_re.match(line)
                if m:
                    candidate = int(m.group(0).split('.')[0].split(')')[0].strip())
                    if candidate > max_fn_seen:
                        flush(current_fn, current_text)
                        current_fn = candidate
                        max_fn_seen = candidate
                        rest = re.sub(r'^\d{1,4}[\.\)]\s+', '', line).strip()
                        current_text = [rest] if rest else []
                    elif current_fn is not None:
                        current_text.append(line)
                elif current_fn is not None:
                    current_text.append(line)

    flush(current_fn, current_text)
    return sorted(footnote_map.items())


# ---------------------------------------------------------------------------
# Step 2: Split footnotes into individual citations
# ---------------------------------------------------------------------------

# Split on: semicolons | before "id." | "see also"
_SPLIT_RE = re.compile(
    r'\s*;\s*'                  # semicolon
    r'|\s+(?=id\.)'             # whitespace immediately before "id."
    r'|\s*[Ss]ee\s+[Aa]lso\s*' # "see also" / "See also"
)


def split_citations(footnote_text: str) -> list[str]:
    """Split a footnote string into individual citation strings."""
    parts = _SPLIT_RE.split(footnote_text)
    return [p.strip() for p in parts if p and p.strip()]


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def extract_citations(
    file_bytes: bytes,
    filename: str,
) -> tuple[pd.DataFrame, dict]:
    """
    Extract and split citations from a .docx or .pdf file.

    Returns (df, metadata) where df has columns:
      footnote_num, footnote_text, citation
    """
    ext = filename.rsplit('.', 1)[-1].lower()
    if ext == 'docx':
        footnotes = extract_footnotes_docx(file_bytes)
    elif ext == 'pdf':
        footnotes = extract_footnotes_pdf(file_bytes)
    else:
        raise ValueError(f'Unsupported file type: {ext}')

    rows = []
    for fn_num, fn_text in footnotes:
        citations = split_citations(fn_text)
        for cite in citations:
            rows.append({'footnote_num': fn_num, 'footnote_text': fn_text, 'citation': cite})

    df = pd.DataFrame(rows, columns=['footnote_num', 'footnote_text', 'citation'])
    metadata = {'total_footnotes': len(footnotes)}
    return df, metadata


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def build_excel(df: pd.DataFrame) -> bytes:
    """Build an Excel workbook with one citation per row."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Citations'

    # Header
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
    ws.append(['Footnote #', 'Citation'])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for _, row in df.iterrows():
        ws.append([row['footnote_num'], row['citation']])

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 100

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
